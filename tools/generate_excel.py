#!/usr/bin/env python3
"""Generate Excel workbooks: shared toolkit, dashboard (Workstream B), datasets."""

from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Participants" / "4_Shared_Toolkit"

NAVY = "1B2A4A"
TEAL = "007A87"
WHITE = "FFFFFF"
HIGHLIGHT = "FFF3E0"

START = datetime(2026, 7, 13)


def style_header(ws, row, ncol, fill=TEAL):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_title(ws, cell_ref, text):
    ws[cell_ref] = text
    ws[cell_ref].font = Font(name="Calibri", bold=True, size=16, color=NAVY)


def auto_width(ws, max_col, min_w=10, max_w=48):
    from openpyxl.utils import get_column_letter
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        mx = min_w
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[letter].width = mx


def week_dates(week):
    return [START + timedelta(weeks=week - 1, days=d) for d in range(5)]


def build_master_timeline(wb):
    ws = wb.active
    ws.title = "Master Timeline"
    style_title(ws, "A1", "Master Timeline and Submission Deadlines")
    ws["A2"] = (
        "Programme start: Monday 13 July 2026. "
        "All deadlines 17:00 Indochina Time (ICT) unless noted."
    )
    ws["A2"].font = Font(italic=True, color="666666")

    headers = [
        "ID", "Week", "Date", "Day", "Time", "Task", "Workflow",
        "Owner", "Deliverable", "Submit to", "Format", "Buddy check-in",
    ]
    r = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers))

    tasks = [
        ("W1-01", 1, 0, "17:00", "Programme kickoff and data room access", "All", "Engagement Lead", "-", "Session", "-", ""),
        ("W1-02", 1, 0, "EOD", "Read handbook, index and workflow briefs", "All", "All", "-", "Team drive", "Internal", ""),
        ("W1-03", 1, 1, "17:00", "Submit signed Project Charter", "All", "PM Lead", "Project_Charter", "Buddy + drive", "PDF", ""),
        ("W1-04", 1, 1, "17:00", "Submit populated RACI matrix", "All", "Engagement Lead", "RACI_Matrix", "Buddy + drive", "Excel", ""),
        ("W1-05", 1, 2, "17:00", "Submit problem statement and issue tree", "A", "WF-A Lead", "A1 Issue tree", "Buddy + drive", "PDF", ""),
        ("W1-06", 1, 3, "17:00", "Submit information credibility matrix", "All", "Research Lead", "Credibility matrix", "Buddy + drive", "PDF", ""),
        ("W1-07", 1, 4, "12:00", "Confirm A/B protocol and Request Log owners", "All", "Engagement Lead", "-", "Buddy", "-", "Check-in 1"),
        ("W1-08", 1, 4, "17:00", "Update Master Timeline with named owners", "All", "PM Lead", "This workbook", "Buddy + drive", "Excel", ""),
        ("W2-01", 2, 0, "17:00", "File Analysis Request R-01", "A", "WF-A Lead", "Analysis_Request", "Team drive", "PDF", ""),
        ("W2-02", 2, 0, "-", "Profile all eight dataset tables (DB-1)", "B", "WF-B Lead", "Profiling notes", "Team drive", "Excel", ""),
        ("W2-03", 2, 1, "17:00", "Submit data quality report (B internal)", "B", "WF-B Lead", "B1 Data quality", "Buddy + drive", "PDF", ""),
        ("W2-04", 2, 2, "17:00", "Submit interview synthesis", "A", "Research Lead", "A2 Synthesis", "Buddy + drive", "PDF", ""),
        ("W2-05", 2, 2, "17:00", "Submit empathy and journey maps", "A", "WF-A Lead", "A3 Maps", "Buddy + drive", "PDF", ""),
        ("W2-06", 2, 3, "17:00", "Return Findings Memo FM-01", "B", "WF-B Lead", "Findings_Memo", "Team drive", "PDF", ""),
        ("W2-07", 2, 4, "17:00", "Submit dashboard version 1 (B working file)", "B", "WF-B Lead", "B2 Dashboard", "Buddy + drive", "Excel", "Check-in 2"),
        ("W3-01", 3, 0, "17:00", "File Analysis Requests R-02 and R-03", "A", "WF-A Lead", "Analysis_Request", "Team drive", "PDF", ""),
        ("W3-02", 3, 1, "17:00", "Return FM-02 and FM-03; analysis summary", "B", "WF-B Lead", "B3 + Findings", "Buddy + drive", "Excel+PDF", ""),
        ("W3-03", 3, 3, "17:00", "Submit strategic options (cite Request IDs)", "A", "WF-A Lead", "A4 Options", "Buddy + drive", "PDF", ""),
        ("W3-04", 3, 4, "17:00", "Submit evidence-check memo on options", "B", "WF-B Lead", "B4 Evidence check", "Team drive", "PDF", "Check-in 3"),
        ("W4-01", 4, 2, "17:00", "Submit validation protocol and trade-off reflection", "A", "Research / WF-A", "A5 A6", "Buddy + drive", "PDF", ""),
        ("W4-02", 4, 3, "17:00", "Submit funding scenarios (DB-5)", "B", "Ops Analyst", "B5 Funding", "Buddy + drive", "Excel", ""),
        ("W4-03", 4, 3, "17:00", "Submit recommendation revision log", "All", "Engagement Lead", "A7 Revision log", "Team drive", "PDF", ""),
        ("W4-04", 4, 4, "17:00", "Submit executive summary (draft)", "A", "Engagement Lead", "Executive summary", "Buddy + drive", "PDF", "Check-in 4"),
        ("W5-01", 5, 1, "17:00", "Submit final report (draft for handoff)", "All", "Engagement Lead", "Final report", "Buddy + drive", "PDF", ""),
        ("W5-02", 5, 2, "12:00", "Submit signed Cross-Workflow Handoff Checklist", "All", "PM Lead", "Handoff checklist", "Buddy", "PDF", ""),
        ("W5-03", 5, 3, "09:00", "BOARD PRESENTATION", "All", "All", "Board deck", "Present live", "PPTX", "Check-in 5"),
        ("W5-04", 5, 4, "17:00", "Submit all final deliverables", "All", "PM Lead", "All finals", "Programme portal", "PDF+PPTX+Excel", ""),
        ("W6-01", 6, 0, "17:00", "Submit peer evaluation (individual, confidential)", "Individual", "Each member", "Peer evaluation", "Portal", "PDF", ""),
        ("W6-02", 6, 1, "17:00", "Submit reflection journal (individual)", "Individual", "Each member", "Reflection", "Portal", "PDF", ""),
    ]

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    r = 5
    fill = PatternFill("solid", fgColor=HIGHLIGHT)
    for t in tasks:
        tid, week, day_idx, time, task, ws_name, owner, deliverable, submit, fmt, buddy = t
        d = week_dates(week)[day_idx]
        vals = [tid, f"Week {week}", d.strftime("%d %b %Y"), days[day_idx], time,
                task, ws_name, owner, deliverable, submit, fmt, buddy]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            if "Submit" in task or "BOARD" in task:
                cell.fill = fill
        r += 1
    ws.freeze_panes = "A5"
    auto_width(ws, 12)


def build_time_tracker(wb):
    ws = wb.create_sheet("Time Tracker")
    style_title(ws, "A1", "Weekly Time Tracker")
    ws["A2"] = "Log hours daily. Submit to your Buddy every Friday by 17:00 ICT."
    headers = ["Date", "Team member", "Activity", "Workflow (A / B / All)", "Hours", "Deliverable"]
    hr = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header(ws, hr, len(headers))
    for _ in range(25):
        ws.append([""] * 6)
    auto_width(ws, 6)


def build_raci(wb):
    ws = wb.create_sheet("RACI Matrix")
    style_title(ws, "A1", "RACI Matrix (R = Responsible, A = Accountable, C = Consulted, I = Informed)")
    headers = ["Task / deliverable", "Eng. Lead", "WF-A Lead", "WF-B Lead", "Member 4", "Member 5", "Member 6"]
    tasks = [
        ("Project Charter", "A", "R", "R", "R", "R", "R"),
        ("Information credibility matrix", "A", "R", "C", "R", "R", "C"),
        ("Analysis Requests (A)", "C", "A/R", "C", "R", "I", "I"),
        ("Findings Memos (B)", "C", "C", "A/R", "R", "R", "I"),
        ("Data quality report (B1)", "C", "C", "A/R", "R", "R", "I"),
        ("Dashboard (B2)", "I", "C", "A/R", "R", "R", "I"),
        ("Interview synthesis (A2)", "C", "A/R", "I", "R", "R", "I"),
        ("Strategic options (A4)", "A", "R", "C", "R", "R", "C"),
        ("Trade-off reflection (A5)", "C", "A/R", "C", "R", "I", "I"),
        ("Validation protocol (A6)", "A", "R", "R", "R", "R", "R"),
        ("Funding scenarios (B5)", "C", "I", "A/R", "R", "R", "I"),
        ("Final report", "A", "R", "R", "R", "R", "R"),
        ("Board deck", "A", "R", "R", "R", "R", "R"),
        ("Cross-workflow handoff", "A", "R", "R", "R", "R", "R"),
    ]
    r = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers))
    for row in tasks:
        ws.append(list(row))
    auto_width(ws, 7)


def build_logs(wb):
    for name, headers in [
        ("Assumption Log", ["ID", "Date", "Assumption", "Source", "Risk if wrong", "Owner", "Validated?", "Outcome"]),
        ("Risk Register", ["ID", "Risk", "Likelihood (1-5)", "Impact (1-5)", "Score", "Mitigation", "Owner", "Status"]),
        ("Decision Log", ["ID", "Date", "Decision", "Options considered", "Rationale", "Dissent", "Owner"]),
        ("Version Control", ["File", "Version", "Date", "Author", "Change", "Reviewed by"]),
    ]:
        ws = wb.create_sheet(name)
        style_title(ws, "A1", name)
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        style_header(ws, 3, len(headers))
        auto_width(ws, len(headers))


def build_funding_scenarios(wb):
    ws = wb.create_sheet("Funding Scenarios")
    style_title(ws, "A1", "Funding Scenarios (Task B5 / DB-5)")
    ws["A2"] = (
        "Use audited unit costs from D-12 and VPBank terms from D-03. "
        "State every assumption. Workflow B owns this tab."
    )
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A4"] = "Reference inputs"
    ws["A4"].font = Font(bold=True, color=NAVY)
    for i, (label, val) in enumerate([
        ("Skills Forward cost per completer (VND)", 12800000),
        ("Pathway Digital cost per completer (VND)", 6200000),
        ("VPBank grant per year (VND)", 2666666667),
        ("Current annual completers (approx.)", 2104),
        ("Broad placement rate (D-01 definition)", "71%"),
        ("Formal placement rate (ST-02)", "63%"),
    ], 5):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)
    r = 12
    headers = ["Scenario", "Growth in completers", "Est. total cost (VND)",
               "Funding gap", "Can meet 65% honestly?", "Key risk"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers))
    for s in [("Hold steady", "0%"), ("Moderate growth", "+30%"), ("Aggressive growth", "+80%")]:
        r += 1
        ws.cell(row=r, column=1, value=s[0])
        ws.cell(row=r, column=2, value=s[1])
    auto_width(ws, len(headers))


def build_submissions(wb):
    ws = wb.create_sheet("Submission Checklist")
    style_title(ws, "A1", "Final Submission Checklist - due Friday 14 August 2026, 17:00 ICT")
    headers = ["Deliverable", "Workflow", "Format", "Week", "Submitted?", "File name"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))
    items = [
        ("Project Charter", "All", "PDF", "1"), ("RACI matrix", "All", "Excel", "1"),
        ("Information credibility matrix", "All", "PDF", "1"), ("Issue tree (A1)", "A", "PDF", "1"),
        ("Stakeholder map (A1b)", "A", "PDF", "1"),
        ("Analysis Requests (min. 3)", "A", "PDF", "2-4"),
        ("Data quality report (B1)", "B", "PDF", "2"), ("Findings Memos (min. 3)", "B", "PDF", "2-4"),
        ("Interview synthesis (A2)", "A", "PDF", "2"),
        ("Empathy and journey maps (A3)", "A", "PDF", "2"), ("Dashboard (B2)", "B", "Excel", "2-3"),
        ("Analysis summary (B3)", "B", "PDF", "3"), ("Strategic options (A4)", "A", "PDF", "3"),
        ("Evidence-check memo (B4)", "B", "PDF", "3"), ("Trade-off reflection (A5)", "A", "PDF", "4"),
        ("Validation protocol (A6)", "All", "PDF", "4"),
        ("Funding scenarios (B5)", "B", "Excel", "4"), ("Revision log (A7)", "All", "PDF", "4"),
        ("Executive summary", "All", "PDF", "4-5"), ("Final report", "All", "PDF", "5"),
        ("Board deck", "All", "PPTX + PDF", "5"), ("Handoff checklist (signed)", "All", "PDF", "5"),
        ("Peer evaluation", "Individual", "PDF", "6"), ("Reflection journal", "Individual", "PDF", "6"),
    ]
    for item in items:
        ws.append(list(item) + ["", ""])
    auto_width(ws, 6)


def build_request_log(wb):
    ws = wb.create_sheet("Request Log")
    style_title(ws, "A1", "Analysis Request Log - minimum three completed R/FM pairs before final options")
    ws["A2"] = (
        "Workflow A files requests. Workflow B returns Findings Memos. "
        "Engagement Lead reviews weekly. Do not share the raw dataset with A."
    )
    ws["A2"].font = Font(italic=True, color="666666")
    headers = [
        "Request ID", "Date filed", "Filed by (A)", "Decision question (short)",
        "Urgency", "Status", "Findings Memo ID", "Date returned", "Returned by (B)", "Notes",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(headers))
    for i in range(1, 9):
        ws.cell(row=4 + i, column=1, value=f"R-{i:02d}")
        ws.cell(row=4 + i, column=6, value="Open")
        ws.cell(row=4 + i, column=7, value=f"FM-{i:02d}")
    auto_width(ws, len(headers))


def build_dashboard_workbook(path: Path):
    """Workflow B dashboard and analysis workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    style_title(ws, "A1", "GBF Performance Dashboard - Workflow B")
    instructions = [
        "This workbook is for Workflow B deliverables B2 and B3.",
        "Source data: GBF_Datasets.xlsx in this same folder (Workflow B only).",
        "Do not share this raw workbook or the datasets file with Workflow A.",
        "Export curated charts into Findings Memos when answering Analysis Requests.",
        "",
        "Task DB-2: Build pivot tables on the Dashboard tab.",
        "  - Placement rate by hub, programme and gender",
        "  - Conditional formatting where rate is below 65%",
        "  - Three charts with insight titles (not topic titles)",
        "",
        "Task DB-3: Attendance vs outcomes on Analysis DB3 tab.",
        "Task DB-4: Mentor hours vs completion on Analysis DB4 tab.",
        "",
        "Rules:",
        "  - State your placement definition on the Dashboard tab",
        "  - Log cleaning decisions in the Shared Toolkit Assumption Log",
        "  - Never delete outliers without documenting why",
    ]
    for i, line in enumerate(instructions, 3):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 90

    dash = wb.create_sheet("Dashboard")
    style_title(dash, "A1", "Performance Dashboard")
    dash["A2"] = "Placement definition used: _______________________________________________"
    dash["A3"] = "Sample note: export covers ~520 records; org total 2,847 per D-01."
    dash["A5"] = "Build your pivot tables and charts below."
    dash["A5"].font = Font(bold=True, color=NAVY)

    for name, title in [("Analysis DB3", "Attendance and placement (DB-3)"),
                        ("Analysis DB4", "Mentor hours and completion (DB-4)"),
                        ("Clean Data", "Paste cleaned tables here")]:
        sh = wb.create_sheet(name)
        style_title(sh, "A1", title)
        sh["A3"] = "Your analysis goes here."

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_master_timeline(wb)
    build_time_tracker(wb)
    build_raci(wb)
    build_logs(wb)
    build_funding_scenarios(wb)
    build_request_log(wb)
    build_submissions(wb)
    toolkit = OUT / "GBF_Consulting_Toolkit.xlsx"
    wb.save(toolkit)
    print(f"  Excel: {toolkit.relative_to(ROOT)}")

    dash_path = ROOT / "Participants" / "3_Workflow_B_Operations_and_Analytics" / "GBF_Performance_Dashboard.xlsx"
    build_dashboard_workbook(dash_path)
    print(f"  Excel: {dash_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()